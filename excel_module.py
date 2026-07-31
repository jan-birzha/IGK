# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import Any

from config import (
    DATA_START_ROW, DEFAULT_FILENAME_COL, DEFAULT_TOTAL_COL, EXCEL_EXTENSIONS,
    L6_L7_FIELDS, L_COLUMN, MANUAL_ROW15_COLS, MANUAL_ROW15_FIELDS, R_COLUMN,
    UI, XL_CALCULATION_MANUAL, XL_SHIFT_DOWN, XL_TEXT_FORMAT
)
from utils import (
    bind_entry_paste_shortcuts, collect_files_with_extensions_recursive,
    dialog_parent, prepare_modal_dialog, show_modal_dialog
)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pythoncom
    import win32com.client as win32
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False


# ── Защита от формульной инъекции в Excel (CWE-1236) ───────────────────────
# Строки, начинающиеся с одного из этих символов, Excel может интерпретировать
# как формулу при присвоении через .Value по COM, независимо от NumberFormat.
_FORMULA_LEAD_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_excel_text(value: object) -> object:
    """
    Нейтрализует потенциальную формульную инъекцию: если строка начинается
    с символа, который Excel трактует как признак формулы, добавляет ведущий
    апостроф, форсирующий текстовый тип ячейки.
    Значения, не являющиеся строками (числа, даты), возвращаются как есть.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_LEAD_CHARS):
        return "'" + value
    return value


def _read_range_row_values(block) -> tuple:
    if block is None:
        return ()
    if isinstance(block, (tuple, list)) and block and isinstance(block[0], (tuple, list)):
        return tuple(block[0])
    if isinstance(block, (tuple, list)):
        return tuple(block)
    return (block,)


def _write_column_values(ws, start_row: int, col: int, values: list) -> None:
    if not values:
        return

    top = ws.Cells(start_row, col)
    bottom = ws.Cells(start_row + len(values) - 1, col)
    target_range = ws.Range(top, bottom) if len(values) > 1 else top

    # Эта функция используется и для текстовых данных (имена файлов), и для
    # чисто числовых (порядковая нумерация в столбце A). Форсируем текстовый
    # формат и нейтрализуем ведущие формульные символы только когда среди
    # значений реально есть строки — иначе не ломаем числовое форматирование.
    has_strings = any(isinstance(v, str) for v in values)
    if has_strings:
        try:
            target_range.NumberFormat = XL_TEXT_FORMAT
            target_range.NumberFormatLocal = XL_TEXT_FORMAT
        except Exception:
            pass
        values = [_sanitize_excel_text(v) for v in values]

    if len(values) == 1:
        top.Value = values[0]
        return
    ws.Range(top, bottom).Value = tuple((v,) for v in values)


# msoAutomationSecurityForceDisable — принудительно отключает выполнение ЛЮБЫХ
# макросов (включая legacy Sub Auto_Open, который срабатывает при открытии
# файла независимо от Application.EnableEvents) без уведомления пользователя.
MSO_AUTOMATION_SECURITY_FORCE_DISABLE = 3


def _configure_excel_fast(excel) -> None:
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False

    # КРИТИЧНО: файлы .xlsx/.xlsm/.xls открываются из произвольной папки,
    # выбранной пользователем, и могут содержать вредоносные макросы
    # (CWE-95, Code Injection). DisplayAlerts=False скрывает от пользователя
    # штатное предупреждение Excel про макросы, поэтому отключение макросов
    # ниже не опционально — при сбое обработка должна прерываться, а не
    # продолжаться в "тихом" режиме с ослабленной защитой.
    try:
        excel.AutomationSecurity = MSO_AUTOMATION_SECURITY_FORCE_DISABLE
    except Exception as exc:
        raise RuntimeError(
            "Не удалось принудительно отключить макросы Excel (AutomationSecurity). "
            "Обработка файлов остановлена в целях безопасности."
        ) from exc

    try:
        excel.Calculation = XL_CALCULATION_MANUAL
    except Exception:
        pass
    try:
        excel.Interactive = False
    except Exception:
        pass


def _restore_excel_defaults(excel) -> None:
    try:
        excel.ScreenUpdating = True
        excel.DisplayAlerts = True
        excel.EnableEvents = True
        excel.CutCopyMode = False
        excel.Interactive = True
    except Exception:
        pass


def name_select_folder(parent: tk.Misc | None = None) -> str | None:
    folder_path = filedialog.askdirectory(parent=dialog_parent(parent), title="Выберите папку")
    return folder_path or None


def name_list_filenames_recursive(folder_path: str) -> list[str]:
    file_names: list[str] = []
    for _current_dir, _sub_dirs, files in os.walk(folder_path):
        for file_name in files:
            file_names.append(file_name)
    return file_names


def name_select_target_file(parent: tk.Misc | None = None) -> str | None:
    file_path = filedialog.askopenfilename(
        parent=dialog_parent(parent),
        title="Выберите файл Excel для вставки данных",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")],
    )
    return file_path or None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_filename_column(ws) -> int:
    block = ws.Range(ws.Cells(12, 1), ws.Cells(14, 39)).Value
    if block is None:
        return DEFAULT_FILENAME_COL
    rows = block if isinstance(block[0], (tuple, list)) else (block,)
    for header_row in rows:
        for col_idx, val in enumerate(header_row, start=1):
            if val and "название файла" in _cell_text(val).lower():
                return col_idx
    return DEFAULT_FILENAME_COL


def _find_total_column(ws, search_row: int) -> int:
    block = ws.Range(ws.Cells(search_row, 1), ws.Cells(search_row, 39)).Value
    row_values = _read_range_row_values(block)
    for col_idx, val in enumerate(row_values, start=1):
        if _cell_text(val).lower() == "итого":
            cell = ws.Cells(search_row, col_idx)
            if cell.MergeCells:
                return cell.MergeArea.Column
            return col_idx
    return DEFAULT_TOTAL_COL


def _unmerge_row_cells(ws, row: int, columns: tuple[str, ...]) -> None:
    for col_letter in columns:
        cell = ws.Range(f"{col_letter}{row}")
        if cell.MergeCells:
            cell.MergeArea.UnMerge()


def _open_workbook(excel, file_path: str):
    abs_path = os.path.normpath(os.path.abspath(file_path))
    if not os.path.isfile(abs_path):
        raise FileNotFoundError(f"Файл не найден: {abs_path}")

    normalized = abs_path.lower()
    for workbook in excel.Workbooks:
        try:
            if os.path.normpath(workbook.FullName).lower() == normalized:
                return workbook
        except Exception:
            continue

    try:
        return excel.Workbooks.Open(abs_path)
    except Exception:
        return excel.Workbooks.Open(
            Filename=abs_path,
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )


def _is_numeric_value(value: object) -> tuple[bool, float]:
    if value is None:
        return False, 0.0
    if isinstance(value, (int, float)):
        return True, float(value)
    text = _cell_text(value).replace(",", ".")
    if not text:
        return False, 0.0
    try:
        return True, float(text)
    except ValueError:
        return False, 0.0


def _cell_target_range(cell):
    if cell.MergeCells:
        return cell.MergeArea
    return cell


def _set_text_format(ws, row: int, col: int) -> None:
    target = _cell_target_range(ws.Cells(row, col))
    target.NumberFormat = XL_TEXT_FORMAT
    try:
        target.NumberFormatLocal = XL_TEXT_FORMAT
    except Exception:
        pass


def _ensure_manual_row_text_format(ws, row: int = DATA_START_ROW) -> None:
    for col in MANUAL_ROW15_COLS:
        _set_text_format(ws, row, col)


def _set_cell_display_value(ws, row: int, col: int, value: str) -> None:
    if not value:
        return
    cell = ws.Cells(row, col)
    target = _cell_target_range(cell)
    target.NumberFormat = XL_TEXT_FORMAT
    try:
        target.NumberFormatLocal = XL_TEXT_FORMAT
    except Exception:
        pass
    # Доп. слой защиты (defense in depth) поверх NumberFormat="@": COM может
    # трактовать строку, начинающуюся с "=", "+" и т.п., как формулу даже
    # при текстовом формате ячейки (CWE-1236, Formula Injection).
    target.Cells(1, 1).Value = _sanitize_excel_text(value)


def _read_manual_row_values(ws, row: int = DATA_START_ROW) -> dict[int, str]:
    block = ws.Range(ws.Cells(row, 2), ws.Cells(row, 6)).Value
    cells = _read_range_row_values(block)
    return {
        2: _cell_text(cells[0]) if len(cells) > 0 else "",
        3: _cell_text(cells[1]) if len(cells) > 1 else "",
        5: _cell_text(cells[3]) if len(cells) > 3 else "",
    }


def _read_manual_row15_openpyxl(path: str) -> dict[int, str] | None:
    if not HAS_OPENPYXL:
        return None
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        def val(coord: str) -> str:
            return _cell_text(ws[coord].value)

        data = {2: val("B15"), 3: val("C15"), 5: val("E15")}
        wb.close()
        return data
    except Exception:
        return None


def _apply_manual_row_entries(ws, row: int, entries: dict[int, str]) -> None:
    for col, value in entries.items():
        if value:
            _set_cell_display_value(ws, row, col, value)


def _apply_l6_l7(ws, l6_value: str, l7_value: str) -> None:
    if l6_value:
        _set_cell_display_value(ws, 6, L_COLUMN, l6_value)
    if l7_value:
        _set_cell_display_value(ws, 7, L_COLUMN, l7_value)


def name_ask_manual_combined(
        parent: tk.Misc | None,
        title: str,
        fields: tuple[tuple[int, str, str], ...],
        include_l6l7: bool,
) -> dict:
    result: dict[Any, str] = {col: "" for col, _, _ in fields}
    if include_l6l7:
        result["L6"] = ""
        result["L7"] = ""

    if not fields and not include_l6l7:
        return result

    dialog = prepare_modal_dialog(parent, title)

    ttk.Label(
        dialog,
        text="Заполните нужные поля (пустое поле — оставить без изменений):",
        background=UI["bg_main"],
        foreground=UI["text"],
        font=UI["font_card_desc"],
    ).pack(padx=20, pady=(16, 12))

    entries: dict[Any, tk.Entry] = {}

    for col, cell_label, prompt in fields:
        frame = tk.Frame(dialog, bg=UI["bg_main"])
        frame.pack(fill=tk.X, padx=20, pady=4)
        tk.Label(
            frame,
            text=f"{cell_label}:",
            background=UI["bg_main"],
            foreground=UI["text"],
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w")
        tk.Label(
            frame,
            text=prompt,
            background=UI["bg_main"],
            foreground=UI["muted"],
            font=UI["font_footer"],
            wraplength=400,
            justify=tk.LEFT
        ).pack(anchor="w")
        entry = tk.Entry(frame, width=50, exportselection=False, font=("Segoe UI", 10), bg=UI["card"], fg=UI["text"],
                         insertbackground=UI["text"], highlightthickness=1, highlightcolor=UI["accent_primary"],
                         highlightbackground=UI["border"], relief=tk.FLAT)
        entry.pack(fill=tk.X, pady=(4, 0))
        bind_entry_paste_shortcuts(dialog, entry)
        entries[col] = entry

    if include_l6l7:
        if fields:
            sep = tk.Frame(dialog, height=1, bg=UI["border"])
            sep.pack(fill=tk.X, padx=20, pady=(8, 4))
        for key, cell_label, prompt in L6_L7_FIELDS:
            frame = tk.Frame(dialog, bg=UI["bg_main"])
            frame.pack(fill=tk.X, padx=20, pady=4)
            tk.Label(
                frame,
                text=f"{cell_label}:",
                background=UI["bg_main"],
                foreground=UI["text"],
                font=("Segoe UI", 9, "bold"),
            ).pack(anchor="w")
            tk.Label(
                frame,
                text=prompt,
                background=UI["bg_main"],
                foreground=UI["muted"],
                font=UI["font_footer"],
                wraplength=400,
                justify=tk.LEFT
            ).pack(anchor="w")
            entry = tk.Entry(frame, width=50, exportselection=False, font=("Segoe UI", 10), bg=UI["card"],
                             fg=UI["text"], insertbackground=UI["text"], highlightthickness=1,
                             highlightcolor=UI["accent_primary"], highlightbackground=UI["border"], relief=tk.FLAT)
            entry.pack(fill=tk.X, pady=(4, 0))
            bind_entry_paste_shortcuts(dialog, entry)
            entries[key] = entry

    if entries:
        next(iter(entries.values())).focus_set()

    btn_frame = tk.Frame(dialog, bg=UI["bg_main"])
    btn_frame.pack(pady=(16, 16))

    def on_ok() -> None:
        for key, entry in entries.items():
            result[key] = entry.get().strip()
        dialog.destroy()

    ok_btn = tk.Button(btn_frame, text="OK", command=on_ok, bg=UI["accent_primary"], fg="#ffffff", relief=tk.FLAT,
                       width=10, font=("Segoe UI", 9, "bold"), activebackground=UI["accent_hover"],
                       activeforeground="#ffffff")
    ok_btn.pack(side=tk.LEFT, padx=6)

    dialog.bind("<Return>", lambda _e: on_ok())
    dialog.bind("<Escape>", lambda _e: on_cancel())

    show_modal_dialog(dialog)
    dialog.wait_window()
    return result


def name_process_workbook(target_path: str, file_names: list[str], parent: tk.Misc | None = None) -> None:
    if not HAS_WIN32:
        messagebox.showerror(
            "Ошибка",
            "Для работы с Excel требуется Windows, Microsoft Excel и pywin32.\nУстановите: pip install pywin32",
            parent=dialog_parent(parent),
        )
        return

    data_count = len(file_names)
    proceed = messagebox.askyesno(
        "Подтверждение",
        f"Будет вставлено {data_count} имён файлов в выбранный файл.\nПродолжить?",
        parent=dialog_parent(parent),
    )
    if not proceed:
        return

    start_total = time.perf_counter()
    timings: list[str] = []

    t0 = time.perf_counter()
    pre_read = _read_manual_row15_openpyxl(target_path)
    manual_entries: dict[int, str] = {}
    l6_value = ""
    l7_value = ""
    if pre_read is not None:
        empty_fields = tuple(
            (col, label, prompt) for col, label, prompt in MANUAL_ROW15_FIELDS if not pre_read.get(col)
        )
        entered = name_ask_manual_combined(parent, "Ручной ввод", empty_fields, True)
        manual_entries = {c: v for c, v in entered.items() if c in (2, 3, 5)}
        l6_value = entered.get("L6", "")
        l7_value = entered.get("L7", "")
        timings.append(f"Ручной ввод: {time.perf_counter() - t0:.2f}s")

    excel = None
    workbook = None
    pythoncom.CoInitialize()

    try:
        t0 = time.perf_counter()
        excel = win32.DispatchEx("Excel.Application")
        _configure_excel_fast(excel)
        timings.append(f"Запуск Excel: {time.perf_counter() - t0:.2f}s")

        t0 = time.perf_counter()
        workbook = _open_workbook(excel, target_path)
        ws = workbook.Worksheets(1)
        timings.append(f"Открытие файла: {time.perf_counter() - t0:.2f}s")

        if pre_read is not None:
            _apply_manual_row_entries(ws, DATA_START_ROW, manual_entries)
            _apply_l6_l7(ws, l6_value, l7_value)
        else:
            t0 = time.perf_counter()
            current = _read_manual_row_values(ws, DATA_START_ROW)
            empty_fields = tuple(
                (col, label, prompt) for col, label, prompt in MANUAL_ROW15_FIELDS if not current.get(col)
            )
            entered = name_ask_manual_combined(parent, "Ручной ввод", empty_fields, True)
            manual_entries = {c: v for c, v in entered.items() if c in (2, 3, 5)}
            _apply_manual_row_entries(ws, DATA_START_ROW, manual_entries)
            _apply_l6_l7(ws, entered.get("L6", ""), entered.get("L7", ""))
            timings.append(f"Ручной ввод: {time.perf_counter() - t0:.2f}s")

        _ensure_manual_row_text_format(ws, DATA_START_ROW)

        t0 = time.perf_counter()
        filename_col = _find_filename_column(ws)
        total_search_row = DATA_START_ROW + 1
        total_col = _find_total_column(ws, total_search_row)
        total_row = DATA_START_ROW + data_count

        cell_a15_value = ws.Cells(DATA_START_ROW, 1).Value
        is_numeric, base_value = _is_numeric_value(cell_a15_value)

        _unmerge_row_cells(ws, 18, ("A",))
        _unmerge_row_cells(ws, 19, ("A",))

        if data_count >= 2:
            insert_count = data_count - 1
            source_row = ws.Rows(DATA_START_ROW)
            target_rows = ws.Rows(f"{DATA_START_ROW + 1}:{DATA_START_ROW + insert_count}")
            source_row.Copy()
            target_rows.Insert(Shift=XL_SHIFT_DOWN)
            excel.CutCopyMode = False

        _write_column_values(ws, DATA_START_ROW, filename_col, file_names)

        if 1 <= data_count <= 5:
            total_cell = ws.Cells(total_row, total_col)
            if total_cell.MergeCells:
                total_cell.MergeArea.Cells(1, 1).Value = "Итого"
            else:
                total_cell.Value = "Итого"

        if data_count >= 2 and is_numeric:
            col_a_values = [base_value + i for i in range(1, data_count)]
            _write_column_values(ws, DATA_START_ROW + 1, 1, col_a_values)

        timings.append(f"Обработка данных: {time.perf_counter() - t0:.2f}s")

        t0 = time.perf_counter()
        workbook.Save()
        timings.append(f"Сохранение: {time.perf_counter() - t0:.2f}s")

        elapsed_total = time.perf_counter() - start_total
        timing_report = "\n".join(f"  • {line}" for line in timings)
        messagebox.showinfo(
            "Готово",
            f"Данные успешно перенесены!\n"
            f"Файл: {os.path.basename(target_path)}\n"
            f"Строк перенесено: {data_count}\n\n"
            f"Время выполнения: {elapsed_total:.2f}s\n"
            f"{timing_report}",
            parent=dialog_parent(parent),
        )
    except Exception as exc:
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n{exc}", parent=dialog_parent(parent))
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                _restore_excel_defaults(excel)
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _apply_manual_entries_to_workbook_array(
        ws,
        manual_entries: dict[int, str],
        l6_value: str,
        l7_value: str,
) -> None:
    row = DATA_START_ROW
    while True:
        r_value = ws.Cells(row, R_COLUMN).Value
        if _cell_text(r_value) == "":
            break

        current = _read_manual_row_values(ws, row)
        to_apply = {
            col: value
            for col, value in manual_entries.items()
            if value and not current.get(col)
        }
        if to_apply:
            _apply_manual_row_entries(ws, row, to_apply)
            _ensure_manual_row_text_format(ws, row)

        row += 1
    _apply_l6_l7(ws, l6_value, l7_value)


def run_name_array(parent: tk.Misc | None = None) -> None:
    if not HAS_WIN32:
        messagebox.showerror(
            "Ошибка",
            "Для работы с Excel требуется Windows, Microsoft Excel и pywin32.\nУстановите: pip install pywin32",
            parent=dialog_parent(parent),
        )
        return

    folder_path = name_select_folder(parent)
    if not folder_path:
        return

    excel_files = collect_files_with_extensions_recursive(folder_path, EXCEL_EXTENSIONS)
    if not excel_files:
        messagebox.showwarning("Внимание", "В выбранной папке нет файлов Excel.", parent=dialog_parent(parent))
        return

    proceed = messagebox.askyesno(
        "Подтверждение",
        f"Найдено файлов Excel: {len(excel_files)}.\nПродолжить обработку?",
        parent=dialog_parent(parent),
    )
    if not proceed:
        return

    entered = name_ask_manual_combined(parent, "Ручной ввод — массив РДО", MANUAL_ROW15_FIELDS, True)
    manual_entries = {c: v for c, v in entered.items() if c in (2, 3, 5) and v}
    l6_value = entered.get("L6", "")
    l7_value = entered.get("L7", "")

    start_total = time.perf_counter()
    pythoncom.CoInitialize()
    excel = None
    processed = 0
    errors: list[str] = []

    try:
        excel = win32.DispatchEx("Excel.Application")
        _configure_excel_fast(excel)

        for path in excel_files:
            workbook = None
            try:
                workbook = _open_workbook(excel, path)
                ws = workbook.Worksheets(1)
                _apply_manual_entries_to_workbook_array(ws, manual_entries, l6_value, l7_value)
                workbook.Save()
                processed += 1
            except Exception as exc:
                errors.append(f"{os.path.basename(path)}: {exc}")
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
    finally:
        if excel is not None:
            try:
                _restore_excel_defaults(excel)
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    elapsed_total = time.perf_counter() - start_total
    summary = f"\n\nОбщее время обработки: {elapsed_total:.2f}s"

    if errors:
        messagebox.showwarning(
            "Готово с предупреждениями",
            f"Обработано успешно: {processed} из {len(excel_files)}.\n\nОшибки:\n" + "\n".join(errors) + summary,
            parent=dialog_parent(parent),
        )
    else:
        messagebox.showinfo("Готово", f"Обработано файлов Excel: {processed} из {len(excel_files)}." + summary, parent=dialog_parent(parent))


def run_name_pick_folder_and_list(parent: tk.Misc | None, listbox: tk.Listbox) -> None:
    folder_path = name_select_folder(parent)
    if not folder_path:
        return

    file_names = name_list_filenames_recursive(folder_path)
    if not file_names:
        messagebox.showwarning("Внимание", "В выбранной папке нет файлов.", parent=dialog_parent(parent))
        return

    listbox.delete(0, tk.END)
    for name in file_names:
        listbox.insert(tk.END, name)


def run_insert_to_rdo(parent: tk.Misc | None, listbox: tk.Listbox) -> None:
    if not HAS_WIN32:
        messagebox.showerror(
            "Ошибка",
            "Для работы с Excel требуется Windows, Microsoft Excel и pywin32.\nУстановите: pip install pywin32",
            parent=dialog_parent(parent),
        )
        return

    file_names = list(listbox.get(0, tk.END))
    if not file_names:
        messagebox.showwarning(
            "Внимание",
            "Список в разделе «Название файла» пуст. Сначала нажмите «Наименование файла в РДО».",
            parent=dialog_parent(parent),
        )
        return

    target_path = name_select_target_file(parent)
    if not target_path:
        return

    name_process_workbook(target_path, file_names, parent)


# ── Новая логика для работы с ДЗО и XML ───────────────────────────────────────

import xml.etree.ElementTree as ET
from xml.dom import minidom


def dzo_select_folder_and_excel(parent: tk.Misc | None = None) -> tuple[str | None, str | None]:
    """Открывает последовательно диалоги выбора папки с файлами и файла Excel."""
    folder_path = name_select_folder(parent)
    if not folder_path:
        return None, None
    excel_path = name_select_target_file(parent)
    if not excel_path:
        return None, None
    return folder_path, excel_path


import datetime as dt

def _format_excel_date(val: object) -> str:
    """Вспомогательная функция для форматирования даты в ДД.ММ.ГГГГ."""
    if val is None:
        return ""
    if isinstance(val, (dt.datetime, dt.date)):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, float) and val > 1000:
        # Дата в виде float (серийный номер Excel)
        try:
            dt_obj = dt.datetime.fromordinal(dt.datetime(1900, 1, 1).toordinal() + int(val) - 2)
            return dt_obj.strftime("%d.%m.%Y")
        except Exception:
            pass
    text = str(val).strip()
    # Если дата записана строкой формата ГГГГ-ММ-ДД
    if len(text) >= 10 and text[4] == '-' and text[7] == '-':
        parts = text[:10].split('-')
        return f"{parts[2]}.{parts[1]}.{parts[0]}"
    return text


def read_dzo_excel_data(excel_path: str) -> list[tuple[str, str, str]]:
    """
    Читает Excel файл (.xlsx, .xlsm, .xls) начиная со 2-й строки.
    Останавливается, если в столбцах A, B, C встречаются пустые ячейки.
    ИГНОРИРУЕТ строки, если значение в столбце I не начинается строго с '40817'.
    Возвращает список кортежей с форматированными данными: (Col_A_Date, Col_B_Text, Col_D_Num).
    """
    ext = os.path.splitext(excel_path)[1].lower()
    result = []

    # Чтение старого формата .xls через xlrd
    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            raise RuntimeError(
                "Для работы с файлами .xls требуется библиотека xlrd.\n"
                "Установите её командой: pip install xlrd"
            )

        wb = xlrd.open_workbook(excel_path)
        ws = wb.sheet_by_index(0)

        row = 1  # 0-based индекс (строка 2 в Excel)
        while row < ws.nrows:
            val_a = ws.cell_value(row, 0)
            val_b = ws.cell_value(row, 1)
            val_c = ws.cell_value(row, 2)

            str_a = _cell_text(val_a)
            str_b = _cell_text(val_b)
            str_c = _cell_text(val_c)

            if not str_a or not str_b or not str_c:
                break

            val_i = ws.cell_value(row, 8)  # Столбец I (индекс 8)
            # Очищаем значение столбца I от случайных пробелов и кавычек
            str_i = _cell_text(val_i).strip().lstrip("'\"").strip()

            # СТРОГОЕ УСЛОВИЕ: Добавляем строку ТОЛЬКО если номер счета начинается на '40817'
            if str_i.startswith("40817"):
                # Форматирование Столбца A (Дата)
                if ws.cell_type(row, 0) == xlrd.XL_CELL_DATE:
                    dt_tuple = xlrd.xldate_as_tuple(val_a, wb.datemode)
                    formatted_a = f"{dt_tuple[2]:02d}.{dt_tuple[1]:02d}.{dt_tuple[0]}"
                else:
                    formatted_a = _format_excel_date(val_a)

                # Форматирование Столбца B (Текст)
                formatted_b = str_b

                # Форматирование Столбца D (Число)
                val_d = ws.cell_value(row, 3)
                is_num, float_d = _is_numeric_value(val_d)
                formatted_d = f"{float_d:,.2f}".replace(",", " ").replace(".", ",") if is_num else _cell_text(val_d)

                result.append((formatted_a, formatted_b, formatted_d))

            row += 1

        return result

    # Чтение форматов .xlsx / .xlsm через openpyxl
    if not HAS_OPENPYXL:
        raise RuntimeError("Для работы с Excel файлом требуется библиотека openpyxl.")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    row = 2
    while True:
        val_a = ws.cell(row=row, column=1).value
        val_b = ws.cell(row=row, column=2).value
        val_c = ws.cell(row=row, column=3).value

        if val_a is None or val_b is None or val_c is None:
            break

        str_a = _cell_text(val_a)
        str_b = _cell_text(val_b)
        str_c = _cell_text(val_c)

        if not str_a or not str_b or not str_c:
            break

        val_i = ws.cell(row=row, column=9).value
        # Очищаем значение столбца I от случайных пробелов и кавычек
        str_i = _cell_text(val_i).strip().lstrip("'\"").strip()

        # СТРОГОЕ УСЛОВИЕ: Добавляем строку ТОЛЬКО если номер счета начинается на '40817'
        if str_i.startswith("40817"):
            # Форматирование A, B, D
            formatted_a = _format_excel_date(val_a)
            formatted_b = str_b

            val_d = ws.cell(row=row, column=4).value
            is_num, float_d = _is_numeric_value(val_d)
            formatted_d = f"{float_d:,.2f}".replace(",", " ").replace(".", ",") if is_num else _cell_text(val_d)

            result.append((formatted_a, formatted_b, formatted_d))

        row += 1

    wb.close()
    return result


def extract_match_tokens(filename: str) -> list[str]:
    """Разбивает имя файла по символу '_' для сопоставления со столбцом B."""
    name_without_ext = os.path.splitext(filename)[0]
    return [token.strip() for token in name_without_ext.split("_") if token.strip()]


def match_dzo_data(excel_rows: list[tuple[str, str, str]], filenames: list[str]) -> list[tuple[str, str, str, str]]:
    """
    Сопоставляет данные из Excel с именами файлов по совпадению столбца B и части имени файла.
    Возвращает список строк таблицы вида: (A, B, D, Filename)
    """
    matched = []
    for filename in filenames:
        tokens = extract_match_tokens(filename)
        found_row = None
        for row in excel_rows:
            col_b = row[1]
            if col_b in tokens:
                found_row = row
                break
        if found_row:
            matched.append((found_row[0], found_row[1], found_row[2], filename))
        # else:
        #     matched.append(("", "", "", filename))
    return matched


def run_dzo_select_folder_and_process(parent: tk.Misc | None, tree: ttk.Treeview) -> None:
    """Обработчик кнопки 'Выбор папки для РДО' на вкладке ДЗО."""
    folder_path, excel_path = dzo_select_folder_and_excel(parent)
    if not folder_path or not excel_path:
        return

    try:
        filenames = name_list_filenames_recursive(folder_path)
        if not filenames:
            messagebox.showwarning("Внимание", "В выбранной папке нет файлов.", parent=dialog_parent(parent))
            return

        excel_rows = read_dzo_excel_data(excel_path)
        matched_data = match_dzo_data(excel_rows, filenames)

        # Очистка и заполнение таблицы РДО
        for item in tree.get_children():
            tree.delete(item)

        for row in matched_data:
            tree.insert("", tk.END, values=row)

    except Exception as exc:
        messagebox.showerror("Ошибка", f"Не удалось обработать файлы:\n{exc}", parent=dialog_parent(parent))


def run_dzo_insert_to_rdo(parent: tk.Misc | None, tree: ttk.Treeview) -> None:
    """
    Обработчик кнопки 'Перенести список в РДО (Excel)' во вкладке ДЗО.
    Поддерживает:
    - Запросом ручного ввода для B15, C15:D15, E15:F15, L6 и L7.
    - Нумерацию строк в столбце A (1, 2, 3...).
    - Перенос A -> I (9), B -> J (10), D -> K (11), Название файла -> filename_col.
    """
    if not HAS_WIN32:
        messagebox.showerror(
            "Ошибка",
            "Для работы с Excel требуется Windows, Microsoft Excel и pywin32.",
            parent=dialog_parent(parent),
        )
        return

    items = tree.get_children()
    if not items:
        messagebox.showwarning("Внимание", "Раздел РДО пуст. Сначала выберите папку и файл Excel.",
                               parent=dialog_parent(parent))
        return

    target_path = name_select_target_file(parent)
    if not target_path:
        return

    rows_data = [tree.item(item)["values"] for item in items]
    data_count = len(rows_data)

    proceed = messagebox.askyesno(
        "Подтверждение",
        f"Будет перенесено {data_count} строк в целевой Excel файл.\nПродолжить?",
        parent=dialog_parent(parent),
    )
    if not proceed:
        return

    # Предварительное считывание значений строки 15 для ручного ввода
    pre_read = _read_manual_row15_openpyxl(target_path)
    manual_entries: dict[int, str] = {}
    l6_value = ""
    l7_value = ""

    if pre_read is not None:
        empty_fields = tuple(
            (col, label, prompt) for col, label, prompt in MANUAL_ROW15_FIELDS if not pre_read.get(col)
        )
        entered = name_ask_manual_combined(parent, "Ручной ввод ДЗО", empty_fields, True)
        manual_entries = {c: v for c, v in entered.items() if c in (2, 3, 5)}
        l6_value = entered.get("L6", "")
        l7_value = entered.get("L7", "")

    pythoncom.CoInitialize()
    excel = None
    workbook = None

    try:
        excel = win32.DispatchEx("Excel.Application")
        _configure_excel_fast(excel)
        workbook = _open_workbook(excel, target_path)
        ws = workbook.Worksheets(1)

        # Если данные в строке 15 не считывались openpyxl
        if pre_read is None:
            current = _read_manual_row_values(ws, DATA_START_ROW)
            empty_fields = tuple(
                (col, label, prompt) for col, label, prompt in MANUAL_ROW15_FIELDS if not current.get(col)
            )
            entered = name_ask_manual_combined(parent, "Ручной ввод ДЗО", empty_fields, True)
            manual_entries = {c: v for c, v in entered.items() if c in (2, 3, 5)}
            l6_value = entered.get("L6", "")
            l7_value = entered.get("L7", "")

        # Применяем ручной ввод для B15, C15, E15 и ячеек L6/L7
        _apply_manual_row_entries(ws, DATA_START_ROW, manual_entries)
        _apply_l6_l7(ws, l6_value, l7_value)
        _ensure_manual_row_text_format(ws, DATA_START_ROW)

        filename_col = _find_filename_column(ws)
        total_search_row = DATA_START_ROW + 1
        total_col = _find_total_column(ws, total_search_row)
        total_row = DATA_START_ROW + data_count

        # Проверка и установка базового номера в A15
        cell_a15_value = ws.Cells(DATA_START_ROW, 1).Value
        is_numeric, base_value = _is_numeric_value(cell_a15_value)
        if not is_numeric or base_value == 0:
            base_value = 1.0
            ws.Cells(DATA_START_ROW, 1).Value = 1

        _unmerge_row_cells(ws, 18, ("A",))
        _unmerge_row_cells(ws, 19, ("A",))

        # Вставка дополнительного количества строк
        if data_count >= 2:
            insert_count = data_count - 1
            source_row = ws.Rows(DATA_START_ROW)
            target_rows = ws.Rows(f"{DATA_START_ROW + 1}:{DATA_START_ROW + insert_count}")
            source_row.Copy()
            target_rows.Insert(Shift=XL_SHIFT_DOWN)
            excel.CutCopyMode = False

        # Заполнение основных данных ДЗО по столбцам
        for idx, row in enumerate(rows_data):
            r = DATA_START_ROW + idx
            val_a, val_b, val_d, fname = row[0], row[1], row[2], row[3]

            _set_cell_display_value(ws, r, 9, str(val_b))  # Столбец I (Дата)
            _set_cell_display_value(ws, r, 10, str(val_a))  # Столбец J (Текст)

            # --- ЗАПИСЬ СУММЫ В ЧИСЛОВОМ ФОРМАТЕ (Столбец K / 11) ---
            cell_k = ws.Cells(r, 11)
            target_k = _cell_target_range(cell_k)

            # Очищаем строку от пробелов и неразрывных пробелов (\xa0)
            val_d_clean = str(val_d).replace(" ", "").replace("\xa0", "")
            is_num, float_val = _is_numeric_value(val_d_clean)

            if is_num:
                # Сначала записываем числовое значение в левую верхнюю ячейку диапазона
                target_k.Cells(1, 1).Value = float_val
                # Безопасно применяем числовой формат для русской/английской локали Excel
                try:
                    target_k.NumberFormatLocal = "# ##0,00"
                except Exception:
                    try:
                        target_k.NumberFormat = "#,##0.00"
                    except Exception:
                        pass  # Если Excel блокирует смену формата, значение всё равно останется числом
            else:
                target_k.Cells(1, 1).Value = val_d

            _set_cell_display_value(ws, r, filename_col, str(fname))  # Наименование файла

        # Заполнение итоговой строки при малом кол-ве элементов
        if 1 <= data_count <= 5:
            total_cell = ws.Cells(total_row, total_col)
            if total_cell.MergeCells:
                total_cell.MergeArea.Cells(1, 1).Value = "Итого"
            else:
                total_cell.Value = "Итого"

        # Нумерация по порядку (1, 2, 3...) в столбце A для всех созданных строк
        if data_count >= 2:
            col_a_values = [int(base_value + i) for i in range(1, data_count)]
            _write_column_values(ws, DATA_START_ROW + 1, 1, col_a_values)

        workbook.Save()
        messagebox.showinfo("Готово", f"Данные успешно перенесены в РДО!\nПеренесено строк: {data_count}",
                            parent=dialog_parent(parent))

    except Exception as exc:
        messagebox.showerror("Ошибка", f"Произошла ошибка при перенесении данных:\n{exc}",
                             parent=dialog_parent(parent))
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                _restore_excel_defaults(excel)
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


# ==============================================================================
# КОНВЕРТАЦИЯ EXCEL В XML (ОБНОВЛЕННАЯ ВЕРСИЯ 2.0)
# ==============================================================================

def _format_xml_date(val: object) -> str:
    """Преобразует дату из Excel в формат XML: YYYY-MM-DD."""
    if val is None:
        return ""
    if isinstance(val, (dt.datetime, dt.date)):
        return val.strftime("%Y-%m-%d")
    text = str(val).strip()
    # Если дата формата ДД.ММ.ГГГГ
    if len(text) >= 10 and text[2] == '.' and text[5] == '.':
        parts = text[:10].split('.')
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    # Если дата формата ГГГГ-ММ-ДД
    if len(text) >= 10 and text[4] == '-' and text[7] == '-':
        return text[:10]
    return text


def _format_xml_sum(val: object) -> str:
    """Преобразует сумму в числовой формат XML: 0.00."""
    if val is None or str(val).strip() == "":
        return "0.00"
    is_num, float_val = _is_numeric_value(str(val).replace(" ", "").replace("\xa0", ""))
    if is_num:
        return f"{float_val:.2f}"
    return str(val).strip()


def run_convert_excel_to_xml(parent: tk.Misc | None = None) -> None:
    """
    Выполняет конвертацию Excel-файла в XML по выбранному шаблону.
    """
    # 1. Шаг 1: Выбор XML-шаблона
    xml_template_path = filedialog.askopenfilename(
        parent=dialog_parent(parent),
        title="Шаг 1 из 3: Выберите файл шаблона XML",
        filetypes=[("XML Files", "*.xml")],
    )
    if not xml_template_path:
        return

    # 2. Шаг 2: Выбор исходного Excel-файла
    excel_path = filedialog.askopenfilename(
        parent=dialog_parent(parent),
        title="Шаг 2 из 3: Выберите Excel файл с данными",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")],
    )
    if not excel_path:
        return

    # 3. Шаг 3: Выбор пути для сохранения результата
    save_xml_path = filedialog.asksaveasfilename(
        parent=dialog_parent(parent),
        title="Шаг 3 из 3: Укажите место для сохранения готового XML файла",
        defaultextension=".xml",
        filetypes=[("XML Files", "*.xml")],
        initialfile="РДО.xml"
    )
    if not save_xml_path:
        return

    ext = os.path.splitext(excel_path)[1].lower()
    excel_rows = []

    # --- Чтение данных из Excel начиная со строки 15 ---
    try:
        if ext == ".xls":
            try:
                import xlrd
            except ImportError:
                messagebox.showerror(
                    "Ошибка",
                    "Для чтения файлов .xls требуется библиотека xlrd.\nУстановите её: pip install xlrd",
                    parent=dialog_parent(parent),
                )
                return

            wb = xlrd.open_workbook(excel_path)
            ws = wb.sheet_by_index(0)

            for r in range(DATA_START_ROW - 1, ws.nrows):
                val_b = ws.cell_value(r, 1)  # B15 (DocDateDB)
                val_c = ws.cell_value(r, 2)  # C15:D15 (DocNumDB)
                val_e = ws.cell_value(r, 4)  # E15:F15 (ContractID)

                # Проверка конца таблицы (если ключевые поля пустые)
                if not _cell_text(val_b) and not _cell_text(val_c) and not _cell_text(val_e):
                    break

                val_sum_ret_raw = ws.cell_value(r, 11)  # L15 (Столбец L, индекс 11)
                is_num, float_sum_ret = _is_numeric_value(val_sum_ret_raw)

                row_item = {
                    "ContractID": _cell_text(val_e),  # E15:F15
                    "DocDateDB": _format_xml_date(val_b),  # B15
                    "DocNumDB": _cell_text(val_c),  # C15:D15
                    "ViewC": _cell_text(ws.cell_value(r, 6)),  # G15:H15
                    "DocNum": _cell_text(ws.cell_value(r, 8)),  # I15
                    "DocDate": _format_xml_date(ws.cell_value(r, 9)),  # J15
                    "Sum": _format_xml_sum(ws.cell_value(r, 10)),  # K15
                    "SumRet": _format_xml_sum(val_sum_ret_raw),  # L15
                    "SumRetFloat": float_sum_ret if is_num else 0.0,
                    "ScanCopy": _cell_text(ws.cell_value(r, 17))  # R15 (Столбец R, индекс 17)
                }
                excel_rows.append(row_item)

        else:  # .xlsx, .xlsm
            if not HAS_OPENPYXL:
                messagebox.showerror(
                    "Ошибка",
                    "Для работы с Excel файлом требуется библиотека openpyxl.",
                    parent=dialog_parent(parent),
                )
                return

            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active

            r = DATA_START_ROW
            while True:
                val_b = ws.cell(row=r, column=2).value  # B15
                val_c = ws.cell(row=r, column=3).value  # C15:D15
                val_e = ws.cell(row=r, column=5).value  # E15:F15

                if val_b is None and val_c is None and val_e is None:
                    break
                if not _cell_text(val_b) and not _cell_text(val_c) and not _cell_text(val_e):
                    break

                val_sum_ret_raw = ws.cell(row=r, column=12).value  # L15 (Столбец L = 12)
                is_num, float_sum_ret = _is_numeric_value(val_sum_ret_raw)

                row_item = {
                    "ContractID": _cell_text(val_e),  # E15:F15
                    "DocDateDB": _format_xml_date(val_b),  # B15
                    "DocNumDB": _cell_text(val_c),  # C15:D15
                    "ViewC": _cell_text(ws.cell(row=r, column=7).value),  # G15:H15
                    "DocNum": _cell_text(ws.cell(row=r, column=9).value),  # I15
                    "DocDate": _format_xml_date(ws.cell(row=r, column=10).value),  # J15
                    "Sum": _format_xml_sum(ws.cell(row=r, column=11).value),  # K15
                    "SumRet": _format_xml_sum(val_sum_ret_raw),  # L15
                    "SumRetFloat": float_sum_ret if is_num else 0.0,
                    "ScanCopy": _cell_text(ws.cell(row=r, column=18).value)  # R15 (Столбец R = 18)
                }
                excel_rows.append(row_item)
                r += 1

            wb.close()

    except Exception as exc:
        messagebox.showerror("Ошибка", f"Не удалось прочитать данные из Excel:\n{exc}", parent=dialog_parent(parent))
        return

    if not excel_rows:
        messagebox.showwarning("Внимание", "В выбранном Excel-файле нет данных начиная со строки 15.",
                               parent=dialog_parent(parent))
        return

    # --- Чтение шаблона ---
    try:
        with open(xml_template_path, "r", encoding="utf-8") as f:
            template_content = f.read()
    except UnicodeDecodeError:
        with open(xml_template_path, "r", encoding="windows-1251") as f:
            template_content = f.read()
    except Exception as exc:
        messagebox.showerror("Ошибка", f"Не удалось прочитать шаблон XML:\n{exc}", parent=dialog_parent(parent))
        return

    # 1. Доработка: Сохранение оригинальной строки <TranscriptPP_PayPurpose> из шаблона
    import re
    from xml.sax.saxutils import escape as _xml_escape, quoteattr as _xml_quoteattr

    def _xml_text(value: object) -> str:
        """Экранирует спецсимволы XML (<, >, &) для текстовых узлов (CWE-91)."""
        return _xml_escape(str(value) if value is not None else "")

    def _xml_attr(value: object) -> str:
        """Экранирует и оборачивает в кавычки значение атрибута XML (CWE-91)."""
        return _xml_quoteattr(str(value) if value is not None else "")

    pay_purpose_match = re.search(r"<TranscriptPP_PayPurpose>(.*?)</TranscriptPP_PayPurpose>", template_content, re.DOTALL)
    if pay_purpose_match:
        pay_purpose_str = pay_purpose_match.group(0)
    else:
        pay_purpose_str = "<TranscriptPP_PayPurpose>Заявка на кассовый расход</TranscriptPP_PayPurpose>"

    # Поиск конца заголовка (до раздела <AnalyticCodePay>)
    tag_target = "</AnalyticCodePay>"
    pos = template_content.find(tag_target)
    if pos != -1:
        header_part = template_content[:pos + len(tag_target)]
    else:
        tag_target = "<AnalyticCodePay"
        pos = template_content.find(tag_target)
        if pos != -1:
            end_pos = template_content.find(">", pos)
            header_part = template_content[:end_pos + 1]
        else:
            messagebox.showerror("Ошибка", "В XML-шаблоне не найден тег <AnalyticCodePay>.",
                                 parent=dialog_parent(parent))
            return

    # --- Генерация элементов <Tab_ReqODB_D07_ITEM> ---
    xml_items = []
    for idx, row in enumerate(excel_rows, start=1):
        view_c_val = row["ViewC"]
        view_c_str = (
            f'<ViewC code="7">{_xml_text(view_c_val)}</ViewC>'
            if view_c_val else '<ViewC code="7">Платежное поручение</ViewC>'
        )

        # 2. Доработка: ScanCopy включается в конец раздела <Tab_ReqODB_D07_ITEM>
        scan_copy_val = row["ScanCopy"]
        scan_copy_tag = (
            f'\n            <ScanCopy attachFileName={_xml_attr(scan_copy_val)}/>'
            if scan_copy_val else ""
        )

        # Все значения из Excel экранируются через _xml_text/_xml_attr перед
        # вставкой в XML — данные пришли из внешнего файла и не должны
        # ломать структуру документа или внедрять посторонние узлы/атрибуты
        # (CWE-91, XML Injection).
        item_xml = f"""        <Tab_ReqODB_D07_ITEM>
            <ContractID>{_xml_text(row['ContractID'])}</ContractID>
            <NumPP>{idx}</NumPP>
            <DocDateDB>{_xml_text(row['DocDateDB'])}</DocDateDB>
            <DocNumDB>{_xml_text(row['DocNumDB'])}</DocNumDB>
            {view_c_str}
            <DocNum>{_xml_text(row['DocNum'])}</DocNum>
            <DocDate>{_xml_text(row['DocDate'])}</DocDate>
            <Sum>{_xml_text(row['Sum'])}</Sum>
            <SumRet>{_xml_text(row['SumRet'])}</SumRet>
            <SumNDS>0.00</SumNDS>{scan_copy_tag}
        </Tab_ReqODB_D07_ITEM>"""
        xml_items.append(item_xml)

    tab_req_odb_content = "\n" + "\n".join(xml_items) + "\n    "

    # 3. Доработка: Подсчет суммы всех чисел столбца L (SumRet)
    total_kbk_sum = sum(row["SumRetFloat"] for row in excel_rows)
    sum_kbk_total_str = f"{total_kbk_sum:.2f}"

    # Сборка итогового XML
    final_xml = f"""{header_part}
</TSE_Tab0401060_ITEM></TSE_Tab0401060><BPR>false</BPR>
<RegistryDocsBaseAttechmentSix_RDOuseSign>true</RegistryDocsBaseAttechmentSix_RDOuseSign>
    <RegistryDocsBaseAttechmentSix_RDONum>1</RegistryDocsBaseAttechmentSix_RDONum>
    <RegistryDocsBaseAttechmentSix_RDODate>{dt.date.today().strftime("%Y-%m-%d")}</RegistryDocsBaseAttechmentSix_RDODate>
    <Tab_ReqODB_D07>{tab_req_odb_content}</Tab_ReqODB_D07>
{pay_purpose_str}
<SumKBKTotal>{sum_kbk_total_str}</SumKBKTotal>
<SumNDSTotal>0.00</SumNDSTotal>
</self:TSE_0401060_D07>"""

    # --- Сохранение в файл ---
    try:
        with open(save_xml_path, "w", encoding="utf-8") as f:
            f.write(final_xml)
        messagebox.showinfo(
            "Успешно",
            f"Конвертация завершена!\n\nФайл сохранен:\n{save_xml_path}\nОбработано строк: {len(excel_rows)}",
            parent=dialog_parent(parent)
        )
    except Exception as exc:
        messagebox.showerror("Ошибка", f"Не удалось сохранить XML файл:\n{exc}", parent=dialog_parent(parent))